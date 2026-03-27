"""
PA45 バッジ自動送信スクリプト

【操作手順】Claudeが以下の順で実行する:

  STEP 1: 日付一覧確認（送信なし）
    python scripts/send-badges.py --session 4 --scan
    → Forms回答がある日付一覧と件数を表示
    → Claudeがユーザーに「どの日付の回答に送りますか？」と確認

  STEP 2: 送信予定リスト確認（送信なし）
    python scripts/send-badges.py --session 4 --date 2026-04-02 --dry-run
    → 送信先メールアドレス一覧を表示
    → Claudeがユーザーに「この宛先でよいですか？」と確認

  STEP 3: 本番送信（ユーザーのOKが出た後のみ）
    python scripts/send-badges.py --session 4 --date 2026-04-02

安全装置:
  - assets/badges/session-XXX/badge.png が存在しない場合は即終了
  - --date なし かつ --scan なし の場合も即終了（誤送信防止）
  - data/badge-sent/session-XXX.json で送信済みを記録 → 重複送信防止

必要な .env 設定:
  BADGE_FORMS_FILE_ID=<FormsのExcelファイルID（OneDrive上）>
  BADGE_FORMS_FILE_OWNER=ixa_mct@plug136.onmicrosoft.com
  SMTP_USER=<送信元メールアドレス>
  SMTP_PASSWORD=<Outlookアプリパスワード>
  NEXT_CONNPASS_URL=<次回connpassのURL>
  MS_TENANT_ID / MS_CLIENT_ID / MS_CLIENT_SECRET=<既存設定>
"""

import os, sys, json, argparse, smtplib, io
from pathlib import Path
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from datetime import datetime, date
from collections import Counter

# ─── 設定読み込み ─────────────────────────────────
_env_path = Path(__file__).parent.parent / ".env"
if _env_path.exists():
    for _line in _env_path.read_text(encoding="utf-8").splitlines():
        if "=" in _line and not _line.startswith("#"):
            _k, _v = _line.split("=", 1)
            os.environ.setdefault(_k.strip(), _v.strip())

ROOT      = Path(__file__).parent.parent
SMTP_HOST = "smtp.office365.com"
SMTP_PORT = 587

# ─── GraphAPIトークン取得 ─────────────────────────
def get_token():
    import requests
    resp = requests.post(
        f"https://login.microsoftonline.com/{os.environ['MS_TENANT_ID']}/oauth2/v2.0/token",
        data={
            "grant_type":    "client_credentials",
            "client_id":     os.environ["MS_CLIENT_ID"],
            "client_secret": os.environ["MS_CLIENT_SECRET"],
            "scope":         "https://graph.microsoft.com/.default",
        }
    )
    resp.raise_for_status()
    return resp.json()["access_token"]

# ─── FormsのExcelを読み込む（共通処理） ──────────
def load_forms_excel(token: str):
    """Excelを読み込んで全行を返す。返り値: (headers, rows, email_col, ts_col)"""
    import requests
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl が必要です → pip install openpyxl")
        sys.exit(1)

    file_id    = os.environ.get("BADGE_FORMS_FILE_ID", "").strip()
    file_owner = os.environ.get("BADGE_FORMS_FILE_OWNER",
                                "ixa_mct@plug136.onmicrosoft.com").strip()

    if not file_id or file_id.startswith("←"):
        print("ERROR: .env に BADGE_FORMS_FILE_ID が設定されていません")
        sys.exit(1)

    url  = f"https://graph.microsoft.com/v1.0/users/{file_owner}/drive/items/{file_id}/content"
    resp = requests.get(url, headers={"Authorization": f"Bearer {token}"})
    resp.raise_for_status()

    wb      = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws      = wb.active
    headers = [str(cell.value or "").strip() for cell in ws[1]]

    # メールアドレス列を探す
    email_col = None
    for i, h in enumerate(headers):
        if any(kw in h.lower() for kw in ["メール", "mail", "email", "e-mail"]):
            email_col = i
            break

    if email_col is None:
        print("ERROR: メールアドレス列が見つかりません")
        print(f"  ヘッダー: {headers}")
        sys.exit(1)

    ts_col = 0  # 開始時刻（タイムスタンプ）は通常0列目
    rows   = list(ws.iter_rows(min_row=2, values_only=True))

    return headers, rows, email_col, ts_col

# ─── タイムスタンプ → dateに変換 ──────────────────
def to_date(ts) -> date | None:
    if isinstance(ts, datetime):
        return ts.date()
    if isinstance(ts, date):
        return ts
    if ts:
        try:
            return datetime.fromisoformat(str(ts)[:10]).date()
        except ValueError:
            pass
    return None

# ─── STEP1: 日付一覧スキャン ──────────────────────
def cmd_scan(token: str, session_num: int):
    """Forms回答の日付一覧と件数を表示する（送信しない）"""
    print("\n📋 Forms回答の日付一覧を取得中...\n")

    _, rows, email_col, ts_col = load_forms_excel(token)

    date_counter: Counter = Counter()
    for row in rows:
        if not any(row):
            continue
        d = to_date(row[ts_col] if len(row) > ts_col else None)
        email = row[email_col] if len(row) > email_col else None
        if d and email and "@" in str(email):
            date_counter[d] += 1

    already_sent = load_sent_log(session_num)

    print("┌────────────────┬───────────┬─────────────┐")
    print("│ 回答日         │ 件数(有効)│ 送信済み    │")
    print("├────────────────┼───────────┼─────────────┤")
    for d in sorted(date_counter.keys(), reverse=True):
        count = date_counter[d]
        print(f"│ {d}  │ {count:>9} │  {len(already_sent):>4}件記録  │")
    print("└────────────────┴───────────┴─────────────┘")
    print()
    print("→ 次のステップ: どの日付の回答に送るか確認して")
    print("  python scripts/send-badges.py"
          f" --session {session_num} --date YYYY-MM-DD --dry-run")

# ─── STEP2/3: メールアドレス取得 ──────────────────
def get_entries_for_date(token: str, target_date: date) -> list[dict]:
    """指定日付の有効なメールアドレス一覧を返す"""
    _, rows, email_col, ts_col = load_forms_excel(token)

    matched = []
    for row in rows:
        if not any(row):
            continue
        d     = to_date(row[ts_col] if len(row) > ts_col else None)
        email = row[email_col] if len(row) > email_col else None
        if d == target_date and email and "@" in str(email):
            ts = row[ts_col] if len(row) > ts_col else None
            matched.append({
                "email":     str(email).strip(),
                "timestamp": str(ts) if ts else "",
            })
    return matched

# ─── 送信済み記録の読み書き ───────────────────────
def load_sent_log(session_num: int) -> set:
    sent_file = ROOT / "data" / "badge-sent" / f"session-{session_num:03d}.json"
    if sent_file.exists():
        data = json.loads(sent_file.read_text(encoding="utf-8"))
        return set(data.get("sent", []))
    return set()

def save_sent_log(session_num: int, sent_emails: set):
    sent_dir  = ROOT / "data" / "badge-sent"
    sent_dir.mkdir(parents=True, exist_ok=True)
    sent_file = sent_dir / f"session-{session_num:03d}.json"
    existing  = load_sent_log(session_num)
    all_sent  = sorted(existing | sent_emails)
    sent_file.write_text(
        json.dumps(
            {"session": session_num, "sent": all_sent,
             "updated": datetime.now().isoformat()},
            ensure_ascii=False, indent=2
        ),
        encoding="utf-8"
    )

# ─── メール送信 ───────────────────────────────────
def send_email(to_email: str, session_num: int, badge_path: Path):
    smtp_user = os.environ.get("SMTP_USER", "").strip()
    smtp_pass = os.environ.get("SMTP_PASSWORD", "").strip()
    next_url  = os.environ.get("NEXT_CONNPASS_URL",
                               "https://connpass.com/group/powerautomate-create/").strip()

    if not smtp_user or smtp_user.startswith("←"):
        print("ERROR: .env に SMTP_USER が設定されていません")
        sys.exit(1)

    subject   = f"【PA45 第{session_num}回】ご参加ありがとうございました！"
    body_html = f"""\
<p>こんにちは！</p>

<p>本日は <strong>PA45 第{session_num}回</strong> にご参加いただき、ありがとうございました！<br>
今日学んだことを、ぜひ明日の業務でひとつ試してみてください。</p>

<hr>

<p>🏅 <strong>参加バッジを添付しました</strong><br>
XなどのSNSでシェアしていただけると嬉しいです！<br>
ハッシュタグ: <strong>#PA45</strong></p>

<hr>

<p>📅 <strong>次回のPA45はこちら</strong><br>
<a href="{next_url}">{next_url}</a></p>

<p>またお会いしましょう！</p>
<p>— PA45 運営 Haru</p>
"""

    msg = MIMEMultipart()
    msg["From"]    = smtp_user
    msg["To"]      = to_email
    msg["Subject"] = subject
    msg.attach(MIMEText(body_html, "html", "utf-8"))

    ext = badge_path.suffix.lstrip(".")
    with open(badge_path, "rb") as f:
        part = MIMEBase("image", ext)
        part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition",
                        f'attachment; filename="PA45_badge_{session_num:03d}.{ext}"')
        msg.attach(part)

    with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as smtp:
        smtp.starttls()
        smtp.login(smtp_user, smtp_pass)
        smtp.sendmail(smtp_user, to_email, msg.as_string())

# ─── バッジ画像パスを確認 ─────────────────────────
def check_badge(session_num: int) -> Path:
    badge_dir = ROOT / "assets" / "badges" / f"session-{session_num:03d}"
    for ext in ["png", "jpg", "jpeg"]:
        p = badge_dir / f"badge.{ext}"
        if p.exists():
            return p
    print(f"\n❌ バッジ画像が見つかりません！送信を中止します。")
    print(f"   配置してください: {badge_dir}/badge.png")
    sys.exit(1)

# ─── メイン ───────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="PA45 バッジ自動送信")
    parser.add_argument("--session",  type=int, required=True, help="回数（例: 4）")
    parser.add_argument("--scan",     action="store_true",
                        help="日付一覧を表示するだけ（送信しない）")
    parser.add_argument("--date",     type=str, default=None,
                        help="対象日 YYYY-MM-DD（--scan なしの場合は必須）")
    parser.add_argument("--dry-run",  action="store_true",
                        help="送信先リストを表示するだけ（送信しない）")
    args = parser.parse_args()

    session_num = args.session

    # ── STEP1: --scan モード ──────────────────────
    if args.scan:
        badge_path = check_badge(session_num)
        print(f"✅ バッジ画像: {badge_path}")
        token = get_token()
        cmd_scan(token, session_num)
        return

    # ── --date 未指定 → 誤送信防止で終了 ──────────
    if not args.date:
        print("\n⛔ --date が指定されていません。")
        print("   まず --scan で日付一覧を確認し、送る日付を --date で指定してください。")
        print(f"   例: python scripts/send-badges.py --session {session_num} --scan")
        sys.exit(1)

    try:
        target_date = date.fromisoformat(args.date)
    except ValueError:
        print("ERROR: 日付の形式が不正です（例: 2026-04-02）")
        sys.exit(1)

    dry_run = args.dry_run

    # ── バッジ画像確認 ────────────────────────────
    badge_path = check_badge(session_num)
    print(f"\n=== PA45 第{session_num}回 バッジ送信 {'[確認モード]' if dry_run else '[本番送信]'} ===")
    print(f"  対象日  : {target_date}")
    print(f"  バッジ  : {badge_path}")

    # ── 送信済み確認 ──────────────────────────────
    already_sent = load_sent_log(session_num)

    # ── Formsから対象日の回答を取得 ───────────────
    print("\n📋 Forms回答を取得中...")
    token   = get_token()
    entries = get_entries_for_date(token, target_date)

    if not entries:
        print(f"\n⚠️  {target_date} のアンケート回答が0件です。")
        print("   --scan で回答がある日付を確認してください。")
        sys.exit(1)

    # ── 未送信に絞り込み ──────────────────────────
    to_send = [e for e in entries
               if e["email"].lower() not in {s.lower() for s in already_sent}]
    skipped = len(entries) - len(to_send)

    # ── 送信予定リストを表示 ──────────────────────
    print(f"\n{'📤 送信予定リスト（確認）' if dry_run else '📤 送信リスト'}:")
    print(f"  {target_date} の回答: {len(entries)}件")
    if skipped:
        print(f"  送信済みスキップ    : {skipped}件")
    print(f"  今回の送信対象      : {len(to_send)}件")
    print()
    for i, e in enumerate(to_send, 1):
        print(f"  {i:>3}. {e['email']}")

    if not to_send:
        print("\n✅ 全員に送信済みです。")
        return

    if dry_run:
        print("\n─────────────────────────────────────────")
        print("  ※ 確認モード（--dry-run）のため送信していません")
        print("  上記リストで問題なければ --dry-run を外して本番送信してください")
        return

    # ── 本番送信 ──────────────────────────────────
    print("\n📨 送信中...")
    newly_sent = set()
    errors     = []

    for i, entry in enumerate(to_send, 1):
        email = entry["email"]
        try:
            send_email(email, session_num, badge_path)
            print(f"  [{i}/{len(to_send)}] ✅ {email}")
            newly_sent.add(email)
        except Exception as e:
            print(f"  [{i}/{len(to_send)}] ❌ {email} → {e}")
            errors.append(email)

    # ── 送信済み記録を保存 ────────────────────────
    if newly_sent:
        save_sent_log(session_num, newly_sent)

    # ── 結果サマリー ─────────────────────────────
    print(f"\n=== 完了 ===")
    print(f"  送信成功: {len(newly_sent)}件")
    if errors:
        print(f"  送信失敗: {len(errors)}件")
        for e in errors:
            print(f"    ❌ {e}")

if __name__ == "__main__":
    main()
